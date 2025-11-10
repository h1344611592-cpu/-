# 导入依赖的库
import requests
import pandas as pd
import subprocess
from rdkit import Chem
from rdkit.Chem import Draw
from tqdm import tqdm
import datetime
import os
import re
import shutil
import time
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


# 自定义函数工具：
"""
目录：
    - makePubChemURL
    - log
    - getPubChemInfo
    - default_process
    - clean_chemical_name
    - replace_greek_letters
    - chinese_to_english_punctuation
    - runSingleSMILES
    - readPredictionOutput
    - appendPrediction
    - save_structure_image
    - append_structure_images
    - insert_images_to_excel
"""
def makePubChemURL(identifier, qtype, output, input=None, 
                   pre="https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound"):
    """
    生成符合PubChem格式的url。
    """
    if input:
        url_x = f"{pre}/{qtype}/{identifier}/{input}/{output.upper()}"
    else:
        url_x = f"{pre}/{qtype}/{identifier}/{output.upper()}"
    
    return url_x


def log(msg, suppress_state=False):
    """
    打印系统消息
    """
    if not suppress_state:
        print(msg)


def getPubChemInfo(url, output_file_name=None, ext="sdf", write_to_folder=True, 
                   get_smiles=False, suppress_state=False, keep_file=True):
    """
    通过url和requests包从PubChem抓取分子信息。
    """
    try:
        response_x = requests.get(url)
        log(f"Status code: {response_x.status_code}", suppress_state=suppress_state)

        if response_x.status_code == 200:
            if write_to_folder:
                if output_file_name is None:
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_file_name = f"sdf_{timestamp}"

                out_file_path = f"data/{output_file_name}.{ext.lower()}"
                with open(out_file_path, 'w') as f:
                    f.write(response_x.text)

                log(f"{output_file_name}.{ext.lower()} is generated to the path：{out_file_path}", suppress_state=suppress_state)

                if get_smiles:
                    mols = [mol for mol in Chem.SDMolSupplier(out_file_path) if mol is not None]
                    smiles = Chem.MolToSmiles(mols[0])

                    if not keep_file and os.path.exists(out_file_path):
                        os.remove(out_file_path)
                        log(f"File has been removed：{out_file_path}", suppress_state=suppress_state)

                    return smiles

                if not keep_file and os.path.exists(out_file_path):
                    os.remove(out_file_path)
                    log(f"已移除当前文件：{out_file_path}", suppress_state=suppress_state)

            else:
                if get_smiles:
                    log("No smiles can be returned, the original response of `requests.get(url)` is returned.\n Please set `write_to_folder=True` and `keep_file=False` to get smiles only.", suppress_state=suppress_state)

            return response_x

        elif response_x.status_code == 404:
            log("Error: 404 - No CID found for the given name", suppress_state=suppress_state)
            return None

        else:
            log(f"Error:\n {response_x.content}", suppress_state=suppress_state)
            return None

    except Exception as e:
        log(f"请求失败：{e}", suppress_state=suppress_state)
        return None


def default_process(name, keep_file=True, suppress_state=False):
    """
    通过化合物英文名一步获取smiles的默认流程。通过分别调用两次`makePubChemURL()`和`getPubChemInfo()`实现，省去了手动调用函数和设置参数。
    """
    try:
        # 生成查询 URL
        url = makePubChemURL(name.strip().lower(), "name", "txt", "cids")
        res = getPubChemInfo(url, write_to_folder=False, 
                             suppress_state=suppress_state)
        
        if res is None or not res.text.strip():  # CID 找不到
            log(f"未找到化合物 {name} 的 CID", suppress_state=suppress_state)
            return "NotFound", None
        
        cid_list = res.text.strip().splitlines()
        
        if len(cid_list) == 0:
            cid = "NotFound"
        elif len(cid_list) > 1:
            cid = "MultipleCIDs"  # 待扩展
        else:
            cid = cid_list[0]  # ✅ 唯一 CID
        
        log(f"{name} 的 CID 是 {cid}", suppress_state=suppress_state)
        
        if cid in ["NotFound", "MultipleCIDs"]:
            return cid, None
        
        # 使用 CID 查询 SDF 文件
        url2 = makePubChemURL(cid, "cid", "sdf")
        smiles = getPubChemInfo(url2, f"cid{cid}", "sdf", get_smiles=True, 
                                suppress_state=suppress_state, 
                                keep_file=keep_file)

        if smiles is None:
            log(f"未获取到 {name}（CID: {cid}） 的结构信息", suppress_state=suppress_state)
            return cid, None  # 只找到文献或结构无效

        return cid, smiles  # 返回有效的 CID 或其他需要的结构信息

    except Exception as e:
        log(f"处理 {name} 时发生错误：{e}", suppress_state=suppress_state)
        return "Error"
    
    
def clean_chemical_name(name):
    if not isinstance(name, str):
        return name

    # 删除特殊字符两侧空格
    name = re.sub(r"\s*([′'`,/-])\s*", r"\1", name)
    name = re.sub(r"\s*([\(\)\[\]\{\}])\s*", r"\1", name)

    # 合并多个空格
    name = re.sub(r"\s+", " ", name)

    # 去除首尾空格
    return name.strip()


def replace_greek_letters(text: str) -> str:
    """
    实现字符串中希腊字母向英文表述的替换。内置希腊字母和英文的对照字典。
    """
    greek_map = {
        "α": "alpha",
        "β": "beta",
        "γ": "gamma",
        "δ": "delta",
        "ε": "epsilon",
        "ζ": "zeta",
        "η": "eta",
        "θ": "theta",
        "ι": "iota",
        "κ": "kappa",
        "λ": "lambda",
        "μ": "mu",
        "ν": "nu",
        "ξ": "xi",
        "ο": "omicron",
        "π": "pi",
        "ρ": "rho",
        "σ": "sigma",
        "τ": "tau",
        "υ": "upsilon",
        "φ": "phi",
        "χ": "chi",
        "ψ": "psi",
        "ω": "omega"
    }
    for greek, latin in greek_map.items():
        text = text.replace(greek, latin)
    return text


def chinese_to_english_punctuation(text):
    """
    将中文标点符号替换为英文标点符号
    
    参数:
        text (str): 包含中文标点符号的文本
        
    返回:
        str: 替换后的文本
    """
    # 创建中英文标点符号对应字典
    punctuation_map = {
        '，': ',',  # 中文逗号
        '。': '.',  # 中文句号
        '、': ',',  # 中文顿号
        '；': ';',  # 中文分号
        '：': ':',  # 中文冒号
        '？': '?',  # 中文问号
        '！': '!',  # 中文感叹号
        '“': '"',  # 中文左双引号
        '”': '"',  # 中文右双引号
        '‘': "'",  # 中文左单引号
        '’': "'",  # 中文右单引号
        '（': '(',  # 中文左括号
        '）': ')',  # 中文右括号
        '《': '<',  # 中文左书名号
        '》': '>',  # 中文右书名号
        '【': '[',  # 中文左方括号
        '】': ']',  # 中文右方括号
        '｛': '{',  # 中文左花括号
        '｝': '}',  # 中文右花括号
        '—': '-',   # 中文破折号
        '…': '...', # 中文省略号
    }
    
    # 逐个字符替换
    for cn, en in punctuation_map.items():
        text = text.replace(cn, en)
    
    return text


def runSingleSMILES(smiles, out_dir="outs", env="VirtuousMultiTaste", 
                    script="D:/VirtuousMultiTaste/VirtuousMultiTaste/VirtuousMultiTaste.py", 
                    debug=False):
    """
    从smiles预测单个化合物的味道。通过subprocess包模拟命令行执行VirtuousMultiTaste的脚本命令。
    """
    demand = ["conda", "run", "-n", env, "python", script, "-c", smiles, "-d", out_dir]
    if debug:
        print("运行命令：", ' '.join(demand))
    subprocess.run(demand, check=True)


def readPredictionOutput(filename="predictions.csv", output_dir="outs", 
                         clean_files=False, suppress_state=False):
    """
    读取预测结果。这里是按VirtuousMultiTaste本身创建的文件路径将模型输出的excel表中的预测结果读取成可用的数据框格式，
    可以更改`filename=your_file.csv`和修改文件存储路径`output_dir=your/path/to/file`读取任何路径下的任意.csv文件。
    """
    filepath = os.path.join(output_dir, filename)
    
    if not os.path.exists(filepath):
        log(f"文件不存在：{filepath}", suppress_state=suppress_state)
        return None
    
    df = pd.read_csv(filepath)
    
    if clean_files:
        try:
            shutil.rmtree(output_dir)
            log(f"已删除文件夹：{output_dir}", suppress_state=suppress_state)
        except Exception as e:
            log(f"删除文件夹时出错：{e}", suppress_state=suppress_state)
                
    return df


def appendPrediction(df, smiles_col="SMILES", env="VirtuousMultiTaste", 
                     script_path="D:/VirtuousMultiTaste/VirtuousMultiTaste/VirtuousMultiTaste.py", 
                     output_file="predictions.csv", output_dir="outs", 
                     clean_files=True, suppress_state=False):
    """
    将预测从单一smiles向数据框扩展，实现`runSingleSMILES()`的批量应用，并输出格式化的结果（向原数据框中添加新的列）。
    """
    pred_result, desc_result, failed_smiles = [], [], []
    
    for smiles in tqdm(df[smiles_col], desc="Running predictions"):
        # 跳过空行
        if pd.isna(smiles) or smiles.strip() == "":
            log("跳过空 SMILES", suppress_state=suppress_state)
            pred_result.append(pd.Series(dtype="object"))
            desc_result.append(pd.Series(dtype="object"))
            continue

        # 跑外部脚本生成结果文件
        try:
            runSingleSMILES(smiles, env=env, script=script_path)
        except subprocess.CalledProcessError:
            log(f"失败的 SMILES: {smiles}", suppress_state=suppress_state)
            failed_smiles.append(smiles)
            pred_result.append(pd.Series(dtype="object"))
            desc_result.append(pd.Series(dtype="object"))
            continue

        # 等待文件生成，避免文件未及时写入的问题
        filepath = os.path.join(output_dir, output_file)
        for _ in range(30):  # 最多等3秒
            if os.path.exists(filepath):
                break
            time.sleep(0.1)

        # 读取文件中的预测结果
        prediction = readPredictionOutput(filename=output_file, 
                                          output_dir=output_dir, 
                                          suppress_state=suppress_state)
        
        descriptors = readPredictionOutput(filename="descriptors.csv",
                                           output_dir=output_dir, 
                                           clean_files=clean_files, 
                                           suppress_state=suppress_state)
        
        if prediction is None or prediction.empty:
            pred_result.append(pd.Series(dtype="object"))
        else:
            pred_row = prediction.iloc[0]

            # 删除与原始表格重复的列
            overlapping_cols = set(df.columns) & set(prediction.columns)
            pred_row = pred_row.drop(labels=overlapping_cols, errors="ignore")

            pred_result.append(pred_row)
        
        if descriptors is None or descriptors.empty:
            desc_result.append(pd.Series(dtype="object"))
        else:
            desc_row = descriptors.iloc[0]

            # 删除与原始表格重复的列
            overlapping_cols = set(df.columns) & set(descriptors.columns)
            desc_row = desc_row.drop(labels=overlapping_cols, errors="ignore")
            
            desc_result.append(desc_row)
    
    # 将所有结果合并为一个 DataFrame
    prediction_df = pd.DataFrame(pred_result).reset_index(drop=True)
    descriptors_df = pd.DataFrame(desc_result).reset_index(drop=True)
    
    # 拼接到原始数据上
    df_final = pd.concat([df.reset_index(drop=True), prediction_df, descriptors_df], axis=1)
    
    if failed_smiles:
        print(f"共有 {len(failed_smiles)} 条 SMILES 运行失败：")
        for s in failed_smiles:
            print("-", s)
            
    return df_final


def save_structure_image(smiles, out_path):
    """
    保存分子结构式的图片。通过rdkit.Chem中的Draw模块获取smiles所对应的二维结构式，并将图片保存到指定路径。
    
    参数:
        smiles: str, SMILES字符串
        out_path: str, 图片保存路径
        
    返回:
        bool: 是否成功保存
    """
    if not smiles or pd.isna(smiles) or smiles == "NoStructure":
        return False
        
    try:
        mol = Chem.MolFromSmiles(smiles)
        if not mol:
            return False
            
        img = Draw.MolToImage(mol)
        # 确保目录存在
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        img.save(out_path)
        return True
    except Exception as e:
        print(f"无法生成 {smiles} 的图片: {str(e)}")
        return False

def append_structure_images(df, smiles_col="SMILES", image_dir="structure_imgs"):
    """
    批量保存图片，并将图片路径保存为数据框中的一列。
    
    参数:
        df: pd.DataFrame, 包含SMILES的数据框
        smiles_col: str, SMILES列名
        image_dir: str, 图片保存目录
        
    返回:
        pd.DataFrame: 添加了StructureImage列的数据框
    """
    # 标准化路径分隔符
    image_dir = os.path.normpath(image_dir)
    os.makedirs(image_dir, exist_ok=True)
    
    image_paths = []
    success_count = 0
    
    for idx, row in df.iterrows():
        smiles = row[smiles_col]
        filename = f"mol_{idx + 1}.png"
        path = os.path.join(image_dir, filename)
        
        if save_structure_image(smiles, path):
            image_paths.append(path)
            success_count += 1
        else:
            image_paths.append(None)
    
    print(f"成功生成 {success_count}/{len(df)} 个分子图片")
    df = df.copy()  # 避免SettingWithCopyWarning
    df["StructureImage"] = image_paths
    return df


def insert_images_to_excel(excel_path, image_column="ImagePath", start_row=2, 
                           image_size=(80, 80)):
    """
    向结果文件中插入结构式。按已导出的excel表中记录的图片路径，将对应图片插入到指定列的对应行（只是插入到那个位置，不能自动调整excel中的单元格大小）
    """
    # 打开已导出的 Excel 文件
    wb = load_workbook(excel_path)
    ws = wb.active

    # 找到图片路径所在列（列名转列号）
    col_idx = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == image_column:
            col_idx = i
            break
    if col_idx is None:
        raise ValueError(f"未找到列名：{image_column}")

    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        img_path = cell.value

        if img_path and os.path.exists(img_path):
            try:
                img = XLImage(img_path)
                img.width, img.height = image_size

                # 将图片插入指定单元格的位置
                cell_coord = f"{get_column_letter(col_idx + 3)}{row}"
                ws.add_image(img, cell_coord)
            except Exception as e:
                print(f"插图失败 at row {row}: {e}")

    # 保存修改
    wb.save(excel_path)
    print(f"图片插入完成，已保存：{excel_path}")

