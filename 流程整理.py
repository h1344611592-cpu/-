# %% 准备工作：导入库
import requests
import pandas as pd
import subprocess
from rdkit import Chem
from rdkit.Chem import Draw
from tqdm import tqdm
import datetime
import os
import shutil
import time
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


tqdm.pandas()  # 添加进度条支持

# %% 准备工作：自定义函数
def makePubChemURL(identifier, qtype, output, input=None, pre="https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound"):
    """
    生成符合PubChem格式的url。
    """
    if input:
        url_x = f"{pre}/{qtype}/{identifier}/{input}/{output.upper()}"
    else:
        url_x = f"{pre}/{qtype}/{identifier}/{output.upper()}"
    
    return url_x


def log(msg, suppress_state=False):
    """
    打印系统消息
    """
    if not suppress_state:
        print(msg)


def getPubChemInfo(url, output_file_name=None, ext="sdf", write_to_folder=True, get_smiles=False, suppress_state=False, keep_file=True):
    """
    通过url和requests包从PubChem抓取分子信息。
    """
    try:
        response_x = requests.get(url)
        log(f"Status code: {response_x.status_code}")

        if response_x.status_code == 200:
            if write_to_folder:
                if output_file_name is None:
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_file_name = f"sdf_{timestamp}"

                out_file_path = f"data/{output_file_name}.{ext.lower()}"
                with open(out_file_path, 'w') as f:
                    f.write(response_x.text)

                log(f"{output_file_name}.{ext.lower()} is generated to the path：{out_file_path}")

                if get_smiles:
                    mols = [mol for mol in Chem.SDMolSupplier(out_file_path) if mol is not None]
                    smiles = Chem.MolToSmiles(mols[0])

                    if not keep_file and os.path.exists(out_file_path):
                        os.remove(out_file_path)
                        log(f"File has been removed：{out_file_path}")

                    return smiles

                if not keep_file and os.path.exists(out_file_path):
                    os.remove(out_file_path)
                    log(f"已移除当前文件：{out_file_path}")

            else:
                if get_smiles:
                    log("No smiles can be returned, the original response of `requests.get(url)` is returned.\n Please set `write_to_folder=True` and `keep_file=False` to get smiles only.")

            return response_x

        elif response_x.status_code == 404:
            log("Error: 404 - No CID found for the given name")
            return None

        else:
            log(f"Error:\n {response_x.content}")
            return None

    except Exception as e:
        log(f"请求失败：{e}")
        return None


def default_process(name, keep_file=True, suppress_state=False):
    """
    通过化合物英文名一步获取smiles的默认流程。通过分别调用两次`makePubChemURL()`和`getPubChemInfo()`实现，省去了手动调用函数和设置参数。
    """
    try:
        # 生成查询 URL
        url = makePubChemURL(name.strip().lower(), "name", "txt", "cids")
        res = getPubChemInfo(url, write_to_folder=False, suppress_state=suppress_state)
        
        if res is None or not res.text.strip():  # CID 找不到
            if not suppress_state:
                print(f"未找到化合物 {name} 的 CID")
            return "NotFound", None
        
        cid_list = res.text.strip().splitlines()
        
        if len(cid_list) == 0:
            cid = "NotFound"
        elif len(cid_list) > 1:
            cid = "MultipleCIDs"  # 待扩展
        else:
            cid = cid_list[0]  # ✅ 唯一 CID
        
        if not suppress_state:
            print(f"{name} 的 CID 是 {cid}")
        
        if cid in ["NotFound", "MultipleCIDs"]:
            return cid, None
        
        # 使用 CID 查询 SDF 文件
        url2 = makePubChemURL(cid, "cid", "sdf")
        smiles = getPubChemInfo(url2, f"cid{cid}", "sdf", get_smiles=True, suppress_state=suppress_state, keep_file=keep_file)

        if smiles is None:
            if not suppress_state:
                print(f"未获取到 {name}（CID: {cid}） 的结构信息")
            return cid, None  # 只找到文献或结构无效

        return cid, smiles  # 返回有效的 CID 或其他需要的结构信息

    except Exception as e:
        print(f"处理 {name} 时发生错误：{e}")
        return "Error"
    

def replace_greek_letters(text: str) -> str:
    """
    实现字符串中希腊字母向英文表述的替换。内置希腊字母和英文的对照字典。
    """
    greek_map = {
        "α": "alpha",
        "β": "beta",
        "γ": "gamma",
        "δ": "delta",
        "ε": "epsilon",
        "ζ": "zeta",
        "η": "eta",
        "θ": "theta",
        "ι": "iota",
        "κ": "kappa",
        "λ": "lambda",
        "μ": "mu",
        "ν": "nu",
        "ξ": "xi",
        "ο": "omicron",
        "π": "pi",
        "ρ": "rho",
        "σ": "sigma",
        "τ": "tau",
        "υ": "upsilon",
        "φ": "phi",
        "χ": "chi",
        "ψ": "psi",
        "ω": "omega"
    }
    for greek, latin in greek_map.items():
        text = text.replace(greek, latin)
    return text


def runSingleSMILES(smiles, out_dir="outs", env="VirtuousMultiTaste", script="D:/VirtuousMultiTaste/VirtuousMultiTaste/VirtuousMultiTaste.py", debug=False):
    """
    从smiles预测单个化合物的味道。通过subprocess包模拟命令行执行VirtuousMultiTaste的脚本命令。
    """
    demand = ["conda", "run", "-n", env, "python", script, "-c", smiles, "-d", out_dir]
    if debug:
        print("运行命令：", ' '.join(demand))
    subprocess.run(demand, check=True)


def readPredictionOutput(filename="predictions.csv", output_dir="outs", clean_files=False, suppress_state=False):
    """
    读取预测结果。这里是按VirtuousMultiTaste本身创建的文件路径将模型输出的excel表中的预测结果读取成可用的数据框格式，
    可以更改`filename=your_file.csv`和修改文件存储路径`output_dir=your/path/to/file`读取任何路径下的任意.csv文件。
    """
    filepath = os.path.join(output_dir, filename)
    
    if not os.path.exists(filepath):
        if not suppress_state:
            print(f"文件不存在：{filepath}")
        return None
    
    df = pd.read_csv(filepath)
    
    if clean_files:
        try:
            shutil.rmtree(output_dir)
            if not suppress_state:
                print(f"已删除文件夹：{output_dir}")
        except Exception as e:
            if not suppress_state:
                print(f"删除文件夹时出错：{e}")
                
    return df


def appendPrediction(df, smiles_col="SMILES", env="VirtuousMultiTaste", script_path="D:/VirtuousMultiTaste/VirtuousMultiTaste/VirtuousMultiTaste.py", 
                     output_file="predictions.csv", output_dir="outs", clean_files=True, suppress_state=False):
    """
    将预测从单一smiles向数据框扩展，实现`runSingleSMILES()`的批量应用，并输出格式化的结果（向原数据框中添加新的列）。
    """
    result, failed_smiles = [], []
    
    for smiles in tqdm(df[smiles_col], desc="Running predictions"):
        # 跳过空行
        if pd.isna(smiles) or smiles.strip() == "":
            if not suppress_state:
                print("跳过空 SMILES")
            result.append(pd.Series(dtype="object"))
            continue

        # 跑外部脚本生成结果文件
        try:
            runSingleSMILES(smiles, env=env, script=script_path)
        except subprocess.CalledProcessError:
            if not suppress_state:
                print("失败的 SMILES:", smiles)
            failed_smiles.append(smiles)
            result.append(pd.Series(dtype="object"))
            continue

        # 等待文件生成，避免文件未及时写入的问题
        filepath = os.path.join(output_dir, output_file)
        for _ in range(30):  # 最多等3秒
            if os.path.exists(filepath):
                break
            time.sleep(0.1)

        # 读取文件中的预测结果
        prediction = readPredictionOutput(filename=output_file, output_dir=output_dir, clean_files=clean_files, suppress_state=suppress_state)

        if prediction is None or prediction.empty:
            result.append(pd.Series(dtype="object"))
        else:
            pred_row = prediction.iloc[0]

            # 删除与原始表格重复的列
            overlapping_cols = set(df.columns) & set(prediction.columns)
            pred_row = pred_row.drop(labels=overlapping_cols, errors="ignore")

            result.append(pred_row)
    
    # 将所有结果合并为一个 DataFrame
    prediction_df = pd.DataFrame(result).reset_index(drop=True)
    
    # 拼接到原始数据上
    df_final = pd.concat([df.reset_index(drop=True), prediction_df], axis=1)
    
    if failed_smiles:
        print(f"共有 {len(failed_smiles)} 条 SMILES 运行失败：")
        for s in failed_smiles:
            print("-", s)
            
    return df_final


def save_structure_image(smiles, out_path):
    """
    保存分子结构式的图片。通过rdkit.Chem中的Draw模块获取smiles所对应的二维结构式，并将图片保存到指定路径。
    """
    mol = Chem.MolFromSmiles(smiles)
    if mol:
        img = Draw.MolToImage(mol)
        img.save(out_path)
        
        
def append_structure_images(df, smiles_col="SMILES", image_dir="structure_imgs"):
    """
    批量保存图片，并将图片路径保存为数据框中的一列。
    """
    os.makedirs(image_dir, exist_ok=True)
    image_paths = []

    for idx, smiles in enumerate(df[smiles_col]):
        filename = f"mol_{idx + 1}.png"
        path = os.path.join(image_dir, filename)
        try:
            save_structure_image(smiles, path)
            image_paths.append(path)
        except:
            image_paths.append(None)
    
    df["StructureImage"] = image_paths
    return df


def insert_images_to_excel(excel_path, image_column="ImagePath", start_row=2, image_size=(80, 80)):
    """
    向结果文件中插入结构式。按已导出的excel表中记录的图片路径，将对应图片插入到指定列的对应行（只是插入到那个位置，不能自动调整excel中的单元格大小）
    """
    # 打开已导出的 Excel 文件
    wb = load_workbook(excel_path)
    ws = wb.active

    # 找到图片路径所在列（列名转列号）
    col_idx = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == image_column:
            col_idx = i
            break
    if col_idx is None:
        raise ValueError(f"未找到列名：{image_column}")

    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        img_path = cell.value

        if img_path and os.path.exists(img_path):
            try:
                img = XLImage(img_path)
                img.width, img.height = image_size

                # 将图片插入指定单元格的位置
                cell_coord = f"{get_column_letter(col_idx + 3)}{row}"
                ws.add_image(img, cell_coord)
            except Exception as e:
                print(f"插图失败 at row {row}: {e}")

    # 保存修改
    wb.save(excel_path)
    print(f"图片插入完成，已保存：{excel_path}")
    
    
# %% 主流程：导入化合物英文名文件。输出格式化的文件：英文名，cid，SMILES，预测结果，（分子式图片）
# 从文件读入化合物列表
df_input = pd.read_excel("data/化合物列表.xlsx")
df_input["英文名"] = (
    df_input["英文名"]
    .astype(str)  # 确保是字符串类型（防止是数字）
    .str.strip()  # 去除两端空白符（空格、换行、tab）
    .str.replace(r"\s*-\s*", "-", regex=True)  # 特例处理：去除连字符周围的空格
    .str.replace(r"\s+", " ", regex=True)  # 把内部多个空白（包括换行、多个空格）变成一个空格
)
# 替换英文名中的希腊字母
df_input["英文名"] = df_input["英文名"].apply(replace_greek_letters)
df_input.tail()  # 预览文件

# %%
# 批量获取化合物的CID和smiles分子式
df_input[["CID", "SMILES"]] = df_input["英文名"].progress_apply(
    lambda name: default_process(name, keep_file=False, suppress_state=True)
).apply(pd.Series)

# %%
# 重跑未找到的数据（减少因为网络问题引起的失败）
retry_mask = df_input["CID"].isin(["NotFound", "MultipleCIDs", None]) | df_input["SMILES"].isin(["NoStructure", None])
# 打印需要重试的数量
print(f"准备重试 {retry_mask.sum()} 条未成功的记录。")
# 只对失败项重试
df_retry = df_input.loc[retry_mask].copy()
df_retry[["CID", "SMILES"]] = df_retry["英文名"].progress_apply(
    lambda name: default_process(name, keep_file=False, suppress_state=True)
).apply(pd.Series)

# %%
# 更新数据框
df_input.update(df_retry)

# %%
# 批量预测
df_result = appendPrediction(df_input, suppress_state=True)
df_result

# %%
# 导出文件
df_result.to_excel("data/result.xlsx", index=False)

# %%
# 插入图片
insert_images_to_excel("data/result.xlsx", image_column="StructureImage")

