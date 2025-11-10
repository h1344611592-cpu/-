# %% 导入库
import requests   #发送HTTP请求
import json    #处理JSON格式数据
import pandas as pd   #表格处理数据（读/写excel）
import subprocess    #调用外部程序（canda环境、预测脚本）
from rdkit import Chem  #分子结构分析（SMILES处理、SDF读取）
from rdkit.Chem import Draw  # 生成分子结构图片
from tqdm import tqdm  # 显示进度条
import datetime    # 生成时间戳（用于文件名）
import os   # 文件/文件夹操作
import shutil   # 批量删除文件
import time   # 等待文件生成（避免IO延迟）
from openpyxl import load_workbook  # 操作Excel（插入图片）
from openpyxl.drawing.image import Image as XLImage  # Excel图片处理
from openpyxl.utils import get_column_letter  # 列号转Excel列名（如1→A）
 

tqdm.pandas()  # 添加进度条支持
# %%
# 执行 API 调用并存储响应
u_2244 = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/2244/property/MolecularFormula,SMILES/JSON"

r = requests.get(u_2244)
print(f"Status code: {r.status_code}")

# %%
# 探索数据结构
rd_2244 = r.json()
rf_2244 = 'data/readable_cid2244_data.json'
with open(rf_2244, 'w') as f:
    json.dump(rd_2244, f, indent=4)
    
# %%
# 打开sdf文件
suppl_2244 = Chem.SDMolSupplier('data/sdf/Structure2D_COMPOUND_CID_2244.sdf')
for mol in suppl_2244:
  print(Chem.MolToSmiles(mol))
  
# %%
# 从cid号到分子式
# 1 编辑访问pubchem的url
query_type = "cid"
cid_x = "2244"
url_pre = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound"
url_input = "property/MolecularFormula,SMILES"
url_output = "JSON"  # JSON通用格式返回部分信息
url_x = f"{url_pre}/{query_type}/{cid_x}/{url_input}/{url_output}"

# 直接读取SDF文件（分子的全部信息）
url_sdf = "SDF"
url_x_full = f"{url_pre}/{query_type}/{cid_x}/{url_sdf}"

# %%
# 2 抓取信息
response_x = requests.get(url_x)
response_x_full = requests.get(url_x_full)
for r in [response_x, response_x_full]:
    print(f"Status code: {r.status_code}")

# %%
# 3 rdkit.Chem返回SMILES
# 先将请求的SDF内容写入文件
with open("data/cid2244.sdf", 'w') as f:
    f.write(response_x_full.text)

# 再用Chem打开
for mol in Chem.SDMolSupplier("data/cid2244.sdf"):
  print(Chem.MolToSmiles(mol))
  print(type(Chem.MolToSmiles(mol)))

# %%
# 另一个分子
def makePubChemURL(identifier, qtype, output, input=None, pre="https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound"):
    if input:
        url_x = f"{pre}/{qtype}/{identifier}/{input}/{output.upper()}"
    else:
        url_x = f"{pre}/{qtype}/{identifier}/{output.upper()}"
    
    return url_x
    
def getPubChemInfo(url, output_file_name=None, ext="sdf", write_to_folder=True, get_smiles=False, suppress_state=False, keep_file=True):
    def log(msg):
        if not suppress_state:
            print(msg)

    try:
        response_x = requests.get(url)
        log(f"Status code: {response_x.status_code}")

        if response_x.status_code == 200:
            if write_to_folder:
                if output_file_name is None:
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_file_name = f"sdf_{timestamp}"

                out_file_path = f"data/{output_file_name}.{ext.lower()}"
                with open(out_file_path, 'w') as f:
                    f.write(response_x.text)

                log(f"{output_file_name}.{ext.lower()}已生成至指定路径：{out_file_path}")

                if get_smiles:
                    mols = [mol for mol in Chem.SDMolSupplier(out_file_path) if mol is not None]
                    smiles = Chem.MolToSmiles(mols[0])

                    if not keep_file and os.path.exists(out_file_path):
                        os.remove(out_file_path)
                        log(f"已移除当前文件：{out_file_path}")

                    return smiles

                if not keep_file and os.path.exists(out_file_path):
                    os.remove(out_file_path)
                    log(f"已移除当前文件：{out_file_path}")

            else:
                if get_smiles:
                    log("No smiles can be returned, the original response of `requests.get(url)` is returned.\n Please set `write_to_folder=True` to get smiles only.")

            return response_x

        elif response_x.status_code == 404:
            log("Error: 404 - No CID found for the given name")
            return None

        else:
            log(f"Error:\n {response_x.content}")
            return None

    except Exception as e:
        log(f"请求失败：{e}")
        return None

url_6322 = makePubChemURL(6322, "cid", "sdf")

smi_6322 = getPubChemInfo(url_6322, "cid6322", get_smiles=True)
print(smi_6322)

getPubChemInfo(url_6322, "cid6322-NOT-SUPPOSED-TO-EXIST", "sdf", get_smiles=True, keep_file=False, suppress_state=True)

# %%
# 获取cid号
url_arg = makePubChemURL("Arginine".lower(), "name", "txt", "cids")

res_arg = getPubChemInfo(url_arg, write_to_folder=False)

print(res_arg.text)

url_glucose = makePubChemURL("glucose".lower(), "name", "txt", "cids")

res_glu = getPubChemInfo(url_glucose, write_to_folder=False)

print(res_glu.text)

url_qa = makePubChemURL("Quinic acid".lower(), "name", "txt", "cids")

res_qa = getPubChemInfo(url_qa, write_to_folder=False)

print(res_qa.text)

# %%
# 从英文名一步到写出smiles
def default_process(name, keep_file=True, suppress_state=False):
    try:
        # 生成查询 URL
        url = makePubChemURL(name.strip().lower(), "name", "txt", "cids")
        res = getPubChemInfo(url, write_to_folder=False, suppress_state=suppress_state)
        
        if res is None or not res.text.strip():  # CID 找不到
            if not suppress_state:
                print(f"未找到化合物 {name} 的 CID")
            return "NotFound", None
        
        cid_list = res.text.strip().splitlines()
        
        if len(cid_list) == 0:
            cid = "NotFound"
        elif len(cid_list) > 1:
            cid = "MultipleCIDs"
        else:
            cid = cid_list[0]  # ✅ 唯一 CID
        
        if not suppress_state:
            print(f"{name} 的 CID 是 {cid}")
        
        if cid in ["NotFound", "MultipleCIDs"]:
            return cid, None
        
        # 使用 CID 查询 SDF 文件
        url2 = makePubChemURL(cid, "cid", "sdf")
        smiles = getPubChemInfo(url2, f"cid{cid}", "sdf", get_smiles=True, suppress_state=suppress_state, keep_file=keep_file)

        if smiles is None:
            if not suppress_state:
                print(f"未获取到 {name}（CID: {cid}） 的结构信息")
            return cid, None  # 只找到文献或结构无效

        return cid, smiles  # 返回有效的 CID 或其他需要的结构信息

    except Exception as e:
        print(f"处理 {name} 时发生错误：{e}")
        return "Error"
    

default_process("Quinic acid", keep_file=False, suppress_state=True)
default_process("curcumin")
# %%
# 特殊字符
default_process('Dihydro-α-cyclohexene lactone')
cid1 = default_process("Quinic acid")        # 正常化合物
cid2 = default_process("UnknownChemicalXYZ") # 未知化合物
cid3 = default_process("gallic acid 3-O-beta-D-glucoside")  # 可能仅返回文献

print(f" {cid1},\n {cid2},\n {cid3}")
# %%
# 输出格式化的文件：英文名，cid，SMILES，（分子式图片）...
# 从文件读入化合物列表
df_input = pd.read_excel("data/化合物列表.xlsx")
df_input["英文名"] = (
    df_input["英文名"]
    .astype(str)  # 确保是字符串类型（防止是数字）
    .str.strip()  # 去除两端空白符（空格、换行、tab）
    .str.replace(r"\s*-\s*", "-", regex=True)  # 特例处理：去除连字符周围的空格
    .str.replace(r"\s+", " ", regex=True)  # 把内部多个空白（包括换行、多个空格）变成一个空格
)
# %%
# 处理希腊字母
greek_map = {
    "α": "alpha",
    "β": "beta",
    "γ": "gamma",
    "δ": "delta",
    "ε": "epsilon",
    "ζ": "zeta",
    "η": "eta",
    "θ": "theta",
    "ι": "iota",
    "κ": "kappa",
    "λ": "lambda",
    "μ": "mu",
    "ν": "nu",
    "ξ": "xi",
    "ο": "omicron",
    "π": "pi",
    "ρ": "rho",
    "σ": "sigma",
    "τ": "tau",
    "υ": "upsilon",
    "φ": "phi",
    "χ": "chi",
    "ψ": "psi",
    "ω": "omega"
}

def replace_greek_letters(text: str) -> str:
    for greek, latin in greek_map.items():
        text = text.replace(greek, latin)
    return text

replace_greek_letters("α, a, b")

# %%
# PubChem别名
default_process("Kaempferol-3-O-beta-D-glucoside", keep_file=False)
default_process("Astragalin", keep_file=False)
# %%
df_input["英文名"] = df_input["英文名"].apply(replace_greek_letters)
df_input.tail()

# %%
# 批量获取smiles分子式
df_input[["CID", "SMILES"]] = df_input["英文名"].progress_apply(
    lambda name: default_process(name, keep_file=False, suppress_state=True)
).apply(pd.Series)

# %%
# 重跑未找到数据
retry_mask = df_input["CID"].isin(["NotFound", "MultipleCIDs", None]) | df_input["SMILES"].isin(["NoStructure", None])
# 打印需要重试的数量
print(f"准备重试 {retry_mask.sum()} 条未成功的记录。")
# 只对失败项重试
df_retry = df_input.loc[retry_mask].copy()
df_retry[["CID", "SMILES"]] = df_retry["英文名"].progress_apply(
    lambda name: default_process(name, keep_file=False, suppress_state=True)
).apply(pd.Series)

# %%
# 更新数据框
df_input.update(df_retry)

# %%
# 结果直接用于预测
env_name = "VirtuousMultiTaste"  # Conda 环境名
script_path = "D:/VirtuousMultiTaste/VirtuousMultiTaste/VirtuousMultiTaste.py"

subprocess.run(["conda", "run", "-n", env_name, "python", script_path, "-c", "C[C@@H]1O[C@@H](OC[C@H]2O[C@@H](Oc3c(-c4ccc(O)c(O)c4)oc4cc(O)cc(O)c4c3=O)[C@H](O)[C@@H](O)[C@@H]2O)[C@H](O)[C@H](O)[C@H]1O", "-d" "outs"])
# %%
def runSingleSMILES(smiles, out_dir="outs", env="VirtuousMultiTaste", script="D:/VirtuousMultiTaste/VirtuousMultiTaste/VirtuousMultiTaste.py", debug=False):
    demand = ["conda", "run", "-n", env, "python", script, "-c", smiles, "-d", out_dir]
    if debug:
        print("运行命令：", ' '.join(demand))
    subprocess.run(demand, check=True)

def readPredictionOutput(filename="predictions.csv", output_dir="outs", clean_files=False, suppress_state=False):
    filepath = os.path.join(output_dir, filename)
    
    if not os.path.exists(filepath):
        if not suppress_state:
            print(f"文件不存在：{filepath}")
        return None
    
    df = pd.read_csv(filepath)
    
    if clean_files:
        try:
            shutil.rmtree(output_dir)
            if not suppress_state:
                print(f"已删除文件夹：{output_dir}")
        except Exception as e:
            if not suppress_state:
                print(f"删除文件夹时出错：{e}")
                
    return df

runSingleSMILES("C[C@@H]1O[C@@H](OC[C@H]2O[C@@H](Oc3c(-c4ccc(O)c(O)c4)oc4cc(O)cc(O)c4c3=O)[C@H](O)[C@@H](O)[C@@H]2O)[C@H](O)[C@H](O)[C@H]1O", "test")
readPredictionOutput(output_dir="test")
# %%
# 应用于数据框
def appendPrediction(df, smiles_col="SMILES", env="VirtuousMultiTaste", script_path="D:/VirtuousMultiTaste/VirtuousMultiTaste/VirtuousMultiTaste.py", 
                     output_file="predictions.csv", output_dir="outs", clean_files=True, suppress_state=False):
    result, failed_smiles = [], []
    
    for smiles in tqdm(df[smiles_col], desc="Running predictions"):
        # 跳过空行
        if pd.isna(smiles) or smiles.strip() == "":
            if not suppress_state:
                print("跳过空 SMILES")
            result.append(pd.Series(dtype="object"))
            continue

        # 跑外部脚本生成结果文件
        try:
            runSingleSMILES(smiles, env=env, script=script_path)
        except subprocess.CalledProcessError:
            if not suppress_state:
                print("失败的 SMILES:", smiles)
            failed_smiles.append(smiles)
            result.append(pd.Series(dtype="object"))
            continue

        # 等待文件生成，避免文件未及时写入的问题
        filepath = os.path.join(output_dir, output_file)
        for _ in range(30):  # 最多等3秒
            if os.path.exists(filepath):
                break
            time.sleep(0.1)

        # 读取文件中的预测结果
        prediction = readPredictionOutput(filename=output_file, output_dir=output_dir, clean_files=clean_files, suppress_state=suppress_state)

        if prediction is None or prediction.empty:
            result.append(pd.Series(dtype="object"))
        else:
            pred_row = prediction.iloc[0]

            # 删除与原始表格重复的列
            overlapping_cols = set(df.columns) & set(prediction.columns)
            pred_row = pred_row.drop(labels=overlapping_cols, errors="ignore")

            result.append(pred_row)
    
    # 将所有结果合并为一个 DataFrame
    prediction_df = pd.DataFrame(result).reset_index(drop=True)
    
    # 拼接到原始数据上
    df_final = pd.concat([df.reset_index(drop=True), prediction_df], axis=1)
    
    if failed_smiles:
        print(f"共有 {len(failed_smiles)} 条 SMILES 运行失败：")
        for s in failed_smiles:
            print("-", s)
            
    return df_final

df_test = df_input.iloc[[0]]
df_test_result = appendPrediction(df_test)
print(df_test_result.columns)
df_test_result

# %%
# 批量处理
df_input = pd.concat([
    df_input[df_input["SMILES"] == "CC(C)[C@H](N)C(=O)O"],  # 这个smiles有问题，挪到第一行好像就没问题了...
    df_input[df_input["SMILES"] != "CC(C)[C@H](N)C(=O)O"]
]).reset_index(drop=True)

df_result = appendPrediction(df_input, suppress_state=True)
df_result

# %%
# 拼接分子式图片路径
def save_structure_image(smiles, out_path):
    mol = Chem.MolFromSmiles(smiles)
    if mol:
        img = Draw.MolToImage(mol)
        img.save(out_path)

def append_structure_images(df, smiles_col="SMILES", image_dir="structure_imgs"):
    os.makedirs(image_dir, exist_ok=True)
    image_paths = []

    for idx, smiles in enumerate(df[smiles_col]):
        filename = f"mol_{idx + 1}.png"
        path = os.path.join(image_dir, filename)
        try:
            save_structure_image(smiles, path)
            image_paths.append(path)
        except:
            image_paths.append(None)
    
    df["StructureImage"] = image_paths
    return df

df_result = append_structure_images(df_result)


# %%
# 导出文件
df_input.to_excel("data/output.xlsx", index=False)

df_result.to_excel("data/result.xlsx", index=False)

# %%
# 插入图片
def insert_images_to_excel(excel_path, image_column="ImagePath", start_row=2, image_size=(80, 80)):
    # 打开已导出的 Excel 文件
    wb = load_workbook(excel_path)
    ws = wb.active

    # 找到图片路径所在列（列名转列号）
    col_idx = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == image_column:
            col_idx = i
            break
    if col_idx is None:
        raise ValueError(f"未找到列名：{image_column}")

    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        img_path = cell.value

        if img_path and os.path.exists(img_path):
            try:
                img = XLImage(img_path)
                img.width, img.height = image_size

                # 将图片插入指定单元格
                cell_coord = f"{get_column_letter(col_idx + 3)}{row}"
                ws.add_image(img, cell_coord)
            except Exception as e:
                print(f"插图失败 at row {row}: {e}")

    # 保存修改
    wb.save(excel_path)
    print(f"图片插入完成，已保存：{excel_path}")

insert_images_to_excel("data/result.xlsx", image_column="StructureImage")
